"""Excel parser for SCC's Profit Summary workbooks.

Real-world files look like ``<year> PS SCC[ variant].xlsx`` (e.g. ``2013 PS SCC.xlsx``,
``2015 PS SCC NEW.xlsx``, ``2016 PS SCC 7.27.16.xlsx``) and follow a stacked-monthly-blocks
layout: one block per month (January → December), each starting with a header row
(``JOB # | PROJECT NAME | % COMPL | CONTRACT | COST | PROFIT | %``), followed by
project rows, a monthly-totals row, and (after Jan) a "CUMULATIVE TOT" row.

Parser is deterministic and side-effect-free. It returns Python dicts that the
worker upserts.
"""
from __future__ import annotations

import logging
import pathlib
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTH_TO_Q = {
    "January": "Q1", "February": "Q1", "March": "Q1",
    "April": "Q2", "May": "Q2", "June": "Q2",
    "July": "Q3", "August": "Q3", "September": "Q3",
    "October": "Q4", "November": "Q4", "December": "Q4",
}


@dataclass
class ProjectRow:
    job_no: str
    name: str
    pct_compl: float | None = None
    contract: float | None = None
    cost: float | None = None
    profit: float | None = None
    profit_pct: float | None = None
    invoiced: float | None = None
    pmt_recd: float | None = None
    last_month: str | None = None


@dataclass
class MonthlyRow:
    month: str
    gross_profit: float | None = None
    overhead: float | None = None
    net_profit: float | None = None


@dataclass
class QuarterlyRow:
    quarter: str
    sales: float | None = None
    gross_profit: float | None = None
    gross_pct: float | None = None
    overhead: float | None = None
    overhead_pct: float | None = None
    net_profit: float | None = None
    net_pct: float | None = None


@dataclass
class OverheadRow:
    month: str
    overhead: float | None = None
    computers: float | None = None
    furniture: float | None = None
    total: float | None = None


@dataclass
class TotalsRow:
    sales: float = 0.0
    gross_profit: float = 0.0
    gross_pct: float = 0.0
    overhead: float = 0.0
    overhead_pct: float = 0.0
    net_profit: float = 0.0
    net_pct: float = 0.0


@dataclass
class ParsedSnapshot:
    fiscal_year: int
    projects: list[ProjectRow] = field(default_factory=list)
    monthly: list[MonthlyRow] = field(default_factory=list)
    quarterly: list[QuarterlyRow] = field(default_factory=list)
    overhead_detail: list[OverheadRow] = field(default_factory=list)
    totals: TotalsRow = field(default_factory=TotalsRow)


# --------------------------------------------------------------------------- #
# Cell helpers
# --------------------------------------------------------------------------- #


def to_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        # Excel sometimes leaves NaN cells; treat as missing.
        if f != f:  # NaN check
            return None
        return f
    s = str(v).strip()
    if not s or s in {"-", "—"}:
        return None
    # Sentinels used in older exports
    if s.upper() in {"INVOICED", "PMT RECD"}:
        return None
    # Excel formula errors
    if s.startswith("#") and s.endswith("!"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s2 = s.strip("()").replace("$", "").replace(",", "").strip()
    pct = s2.endswith("%")
    s2 = s2.rstrip("%").strip()
    try:
        n = float(s2)
    except ValueError:
        return None
    if neg:
        n = -n
    if pct:
        n = n / 100.0
    return n


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())


def _year_from_filename(path: pathlib.Path) -> int | None:
    """Extract the fiscal year from a filename like '2013 PS SCC.xlsx' or
    'Profitability_2024.xlsx'. Returns the first 4-digit year in 19xx–20xx range."""
    m = re.search(r"((?:19|20)\d{2})", path.name)
    return int(m.group(1)) if m else None


def _is_template(path: pathlib.Path) -> bool:
    return "template" in path.name.lower()


def _all_rows(path: pathlib.Path) -> list[list[Any]]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    out: list[list[Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            out.append(list(row))
    wb.close()
    return out


# --------------------------------------------------------------------------- #
# Stacked-monthly-blocks parser (the real SCC format)
# --------------------------------------------------------------------------- #


def _is_header_row(row: list[Any]) -> bool:
    """A block header has 'JOB' in col B and 'PROJECT NAME' in col C."""
    if len(row) < 3:
        return False
    b = _norm(row[1])
    c = _norm(row[2])
    return b.startswith("job") and "projectname" in c


def _is_cumulative_row(row: list[Any]) -> bool:
    if len(row) < 3:
        return False
    return "cumulative" in _norm(row[2])


def _is_monthly_totals_row(row: list[Any]) -> bool:
    """A monthly-totals row: ###, JOB#, PROJECT NAME all blank; CONTRACT has a number."""
    if len(row) < 7:
        return False
    a = str(row[0] or "").strip()
    b = str(row[1] or "").strip()
    c = str(row[2] or "").strip()
    if a or b or c:
        return False
    return to_number(row[4]) is not None or to_number(row[6]) is not None


def parse_project_list_blocks(
    rows: list[list[Any]],
) -> tuple[list[ProjectRow], list[MonthlyRow]]:
    """Parse one workbook's worth of rows in the stacked-monthly-blocks layout."""
    projects: list[ProjectRow] = []
    monthly: list[MonthlyRow] = []
    month_idx = -1
    saw_header = False
    in_block_after_totals = False

    for r in rows:
        if not r or all(c in (None, "") for c in r):
            continue

        if _is_header_row(r):
            month_idx += 1
            saw_header = True
            in_block_after_totals = False
            continue

        if not saw_header:
            continue

        if _is_cumulative_row(r):
            continue

        if _is_monthly_totals_row(r):
            if 0 <= month_idx < 12:
                month_name = MONTHS[month_idx]
                gp = to_number(r[6]) if len(r) > 6 else None
                # If a later block reports a different total for the same month,
                # last-write-wins (consistent with the file order).
                existing = next((m for m in monthly if m.month == month_name), None)
                if existing:
                    existing.gross_profit = gp
                else:
                    monthly.append(MonthlyRow(month=month_name, gross_profit=gp))
            in_block_after_totals = True
            continue

        # Project rows: JOB# + PROJECT NAME present; ### usually has the within-block index.
        job_raw = r[1] if len(r) > 1 else None
        name_raw = r[2] if len(r) > 2 else None
        job = str(job_raw or "").strip() if job_raw not in (None, "") else ""
        name = str(name_raw or "").strip() if name_raw not in (None, "") else ""
        if not job or not name:
            continue
        if in_block_after_totals:
            # Stray rows after the totals (e.g. cumulative cash) — skip.
            continue

        try:
            month_name = MONTHS[month_idx] if 0 <= month_idx < 12 else None
            contract = to_number(r[4]) if len(r) > 4 else None
            cost = to_number(r[5]) if len(r) > 5 else None
            profit = to_number(r[6]) if len(r) > 6 else None
            profit_pct = to_number(r[7]) if len(r) > 7 else None
            # Derive missing pieces (matches the original embedded dataset)
            if profit is None and contract is not None and cost is not None:
                profit = contract - cost
            if (
                profit_pct is None
                and profit is not None
                and contract not in (None, 0)
            ):
                profit_pct = profit / contract
            projects.append(
                ProjectRow(
                    job_no=job,
                    name=name,
                    pct_compl=to_number(r[3]) if len(r) > 3 else None,
                    contract=contract,
                    cost=cost,
                    profit=profit,
                    profit_pct=profit_pct,
                    invoiced=to_number(r[14]) if len(r) > 14 else None,
                    pmt_recd=to_number(r[16]) if len(r) > 16 else None,
                    last_month=month_name,
                )
            )
        except Exception as exc:  # pragma: no cover -- defensive
            log.warning("skipping bad project row: %s", exc)

    return projects, monthly


def parse_overhead_detail(rows: list[list[Any]]) -> list[OverheadRow]:
    """The right-hand side of each PS SCC workbook has a monthly overhead block:

        Month     | Overhead | Computers | Furniture | Total
        January   | 25014.98 |          |          | 25014.98
        February  | 26285.30 |          |          | 26285.30
        ...

    Find the header by scanning every row for five consecutive cells matching
    ``[month, overhead, computers, furniture, total]`` (any column offset), then
    read up to 12 months below it.
    """
    out: list[OverheadRow] = []
    for ri, row in enumerate(rows):
        for ci in range(len(row) - 4):
            window = [_norm(row[ci + k]) for k in range(5)]
            if window == ["month", "overhead", "computers", "furniture", "total"]:
                for dr in range(1, 13):
                    if ri + dr >= len(rows):
                        break
                    drow = rows[ri + dr]
                    if ci >= len(drow):
                        break
                    month_cell = drow[ci]
                    if not isinstance(month_cell, str):
                        break
                    name = month_cell.strip()
                    if name not in MONTHS:
                        break
                    out.append(
                        OverheadRow(
                            month=name,
                            overhead=to_number(drow[ci + 1]) if ci + 1 < len(drow) else None,
                            computers=to_number(drow[ci + 2]) if ci + 2 < len(drow) else None,
                            furniture=to_number(drow[ci + 3]) if ci + 3 < len(drow) else None,
                            total=to_number(drow[ci + 4]) if ci + 4 < len(drow) else None,
                        )
                    )
                return out
    return out


# --------------------------------------------------------------------------- #
# Legacy single-header parser (kept for the dev seed / synthetic test fixtures)
# --------------------------------------------------------------------------- #


PROJECT_HEADERS = {
    "job": ("jobno", "job", "job#"),
    "name": ("name", "projectname", "project"),
    "pct_compl": ("pctcompl", "complete", "compl"),
    "contract": ("contract", "contractvalue", "contracttotal"),
    "cost": ("cost", "totalcost"),
    "profit": ("profit", "grossprofit"),
    "profit_pct": ("margin", "profitpct", "profitmargin"),
    "invoiced": ("invoiced",),
    "pmt_recd": ("pmtrecd", "paymentreceived", "paid"),
    "last_month": ("lastmonth", "month"),
}
OVERHEAD_HEADERS = {
    "month": ("month",),
    "overhead": ("overhead",),
    "computers": ("computers",),
    "furniture": ("furniture",),
    "total": ("total",),
}


def _find_header(rows, schema):
    required = {"job", "name"} if schema is PROJECT_HEADERS else {"month"}
    for i, row in enumerate(rows):
        normalized = [_norm(c) for c in row]
        mapping: dict[str, int] = {}
        for field_name, tokens in schema.items():
            for tok in tokens:
                if tok in normalized:
                    mapping[field_name] = normalized.index(tok)
                    break
        if required <= set(mapping.keys()):
            return i, mapping
    return None


def parse_profitability_single_header(rows: list[list[Any]]) -> list[ProjectRow]:
    located = _find_header(rows, PROJECT_HEADERS)
    if not located:
        return []
    start, cols = located
    out: list[ProjectRow] = []
    for r in rows[start + 1 :]:
        if cols.get("job") is None or cols["job"] >= len(r):
            continue
        if cols.get("name") is None or cols["name"] >= len(r):
            continue
        if not r[cols["job"]] or not r[cols["name"]]:
            continue
        try:
            row = ProjectRow(
                job_no=str(r[cols["job"]]).strip(),
                name=str(r[cols["name"]]).strip(),
                pct_compl=to_number(r[cols["pct_compl"]]) if "pct_compl" in cols and cols["pct_compl"] < len(r) else None,
                contract=to_number(r[cols["contract"]]) if "contract" in cols and cols["contract"] < len(r) else None,
                cost=to_number(r[cols["cost"]]) if "cost" in cols and cols["cost"] < len(r) else None,
                profit=to_number(r[cols["profit"]]) if "profit" in cols and cols["profit"] < len(r) else None,
                profit_pct=to_number(r[cols["profit_pct"]]) if "profit_pct" in cols and cols["profit_pct"] < len(r) else None,
                invoiced=to_number(r[cols["invoiced"]]) if "invoiced" in cols and cols["invoiced"] < len(r) else None,
                pmt_recd=to_number(r[cols["pmt_recd"]]) if "pmt_recd" in cols and cols["pmt_recd"] < len(r) else None,
                last_month=str(r[cols["last_month"]]).strip() if "last_month" in cols and cols["last_month"] < len(r) and r[cols["last_month"]] else None,
            )
            if row.profit is None and row.contract is not None and row.cost is not None:
                row.profit = row.contract - row.cost
            if row.profit_pct is None and row.profit is not None and row.contract not in (None, 0):
                row.profit_pct = row.profit / row.contract
            out.append(row)
        except Exception as exc:  # pragma: no cover -- defensive
            log.warning("skipping bad project row: %s", exc)
    return out


def parse_overhead(rows: list[list[Any]]) -> list[OverheadRow]:
    located = _find_header(rows, OVERHEAD_HEADERS)
    if not located:
        return []
    start, cols = located
    out: list[OverheadRow] = []
    for r in rows[start + 1 :]:
        month_v = r[cols["month"]] if "month" in cols and cols["month"] < len(r) else None
        if not month_v:
            continue
        month = str(month_v).strip()
        if month not in MONTHS:
            continue
        oh = to_number(r[cols["overhead"]]) if "overhead" in cols and cols["overhead"] < len(r) else None
        comp = to_number(r[cols["computers"]]) if "computers" in cols and cols["computers"] < len(r) else 0
        furn = to_number(r[cols["furniture"]]) if "furniture" in cols and cols["furniture"] < len(r) else 0
        total = to_number(r[cols["total"]]) if "total" in cols and cols["total"] < len(r) else None
        if total is None:
            total = (oh or 0) + (comp or 0) + (furn or 0)
        out.append(
            OverheadRow(month=month, overhead=oh, computers=comp, furniture=furn, total=total)
        )
    return out


# --------------------------------------------------------------------------- #
# File dispatch + aggregation
# --------------------------------------------------------------------------- #


def parse_workbook(
    path: pathlib.Path,
) -> tuple[list[ProjectRow], list[MonthlyRow], list[OverheadRow]]:
    """Parse one workbook. Tries the SCC stacked-monthly-blocks format first,
    then falls back to the legacy single-header format used by the dev seed."""
    rows = _all_rows(path)
    projects, monthly = parse_project_list_blocks(rows)
    overhead = parse_overhead_detail(rows)
    if projects:
        return projects, monthly, overhead
    return parse_profitability_single_header(rows), [], []


def compute_monthly_from_projects(
    projects: list[ProjectRow],
    overhead: list[OverheadRow],
    block_monthly: list[MonthlyRow] | None = None,
) -> list[MonthlyRow]:
    """Build the 12-month series. Prefer the totals reported in each block when
    available; otherwise sum project profits per ``last_month``."""
    oh_by_month = {o.month: (o.total or 0) for o in overhead}
    block_by_month = {m.month: m for m in (block_monthly or [])}

    gp_by_month: dict[str, float] = {m: 0.0 for m in MONTHS}
    for p in projects:
        if p.last_month and p.last_month in gp_by_month and p.profit is not None:
            gp_by_month[p.last_month] += float(p.profit)

    rows: list[MonthlyRow] = []
    for m in MONTHS:
        block = block_by_month.get(m)
        gp = float(block.gross_profit) if block and block.gross_profit is not None else gp_by_month[m]
        oh = oh_by_month.get(m, 0.0)
        rows.append(MonthlyRow(month=m, gross_profit=gp, overhead=oh, net_profit=gp - oh))
    return rows


def compute_quarterly_from_projects(
    projects: list[ProjectRow], monthly: list[MonthlyRow]
) -> list[QuarterlyRow]:
    monthly_by = {m.month: m for m in monthly}
    sales_by_q: dict[str, float] = {"Q1": 0, "Q2": 0, "Q3": 0, "Q4": 0}
    gp_by_q = {"Q1": 0.0, "Q2": 0.0, "Q3": 0.0, "Q4": 0.0}
    oh_by_q = {"Q1": 0.0, "Q2": 0.0, "Q3": 0.0, "Q4": 0.0}
    for p in projects:
        if p.last_month and p.last_month in MONTH_TO_Q and p.contract is not None:
            sales_by_q[MONTH_TO_Q[p.last_month]] += float(p.contract)
    for m in MONTHS:
        q = MONTH_TO_Q[m]
        mrow = monthly_by.get(m)
        if mrow:
            gp_by_q[q] += float(mrow.gross_profit or 0)
            oh_by_q[q] += float(mrow.overhead or 0)

    rows: list[QuarterlyRow] = []
    for q in ("Q1", "Q2", "Q3", "Q4"):
        sales = sales_by_q[q]
        gp = gp_by_q[q]
        oh = oh_by_q[q]
        net = gp - oh
        rows.append(
            QuarterlyRow(
                quarter=q, sales=sales, gross_profit=gp,
                gross_pct=(gp / sales) if sales else 0,
                overhead=oh, overhead_pct=(oh / sales) if sales else 0,
                net_profit=net, net_pct=(net / sales) if sales else 0,
            )
        )
    return rows


def compute_totals(quarterly: list[QuarterlyRow]) -> TotalsRow:
    sales = sum((q.sales or 0) for q in quarterly)
    gp = sum((q.gross_profit or 0) for q in quarterly)
    oh = sum((q.overhead or 0) for q in quarterly)
    net = gp - oh
    return TotalsRow(
        sales=sales, gross_profit=gp,
        gross_pct=(gp / sales) if sales else 0,
        overhead=oh, overhead_pct=(oh / sales) if sales else 0,
        net_profit=net, net_pct=(net / sales) if sales else 0,
    )


def _dedupe_projects(projects: list[ProjectRow]) -> list[ProjectRow]:
    """Multiple files for the same year (e.g. '2015 PS SCC' + '2015 PS SCC NEW')
    contribute overlapping job rows. Keep the LAST occurrence for each job_no."""
    seen: dict[str, int] = {}
    for i, p in enumerate(projects):
        seen[p.job_no] = i
    return [projects[i] for i in sorted(seen.values())]


def parse_files_for_year(files: list[pathlib.Path], fiscal_year: int) -> ParsedSnapshot:
    """Parse one or more workbooks that all belong to the same fiscal year.

    When multiple files map to the same year (e.g. ``2015 PS SCC.xlsx`` +
    ``2015 PS SCC NEW.xlsx``), they're processed in order of modification time
    (oldest first), so the newest file's data wins on duplicate job numbers.
    """
    files = sorted(files, key=lambda p: p.stat().st_mtime)
    snap = ParsedSnapshot(fiscal_year=fiscal_year)
    block_monthly: list[MonthlyRow] = []
    for path in files:
        if _is_template(path):
            log.info("skipping template file %s", path.name)
            continue
        try:
            projects, monthly, overhead = parse_workbook(path)
        except Exception:
            log.exception("failed to parse %s", path.name)
            continue
        snap.projects.extend(projects)
        # Block-reported monthly totals: later file wins.
        by_m = {m.month: m for m in block_monthly}
        for m in monthly:
            by_m[m.month] = m
        block_monthly = list(by_m.values())
        # Overhead detail: later file wins per month.
        oh_by_m = {o.month: o for o in snap.overhead_detail}
        for o in overhead:
            oh_by_m[o.month] = o
        snap.overhead_detail = list(oh_by_m.values())

    snap.projects = _dedupe_projects(snap.projects)
    snap.monthly = compute_monthly_from_projects(
        snap.projects, snap.overhead_detail, block_monthly=block_monthly
    )
    snap.quarterly = compute_quarterly_from_projects(snap.projects, snap.monthly)
    snap.totals = compute_totals(snap.quarterly)
    return snap


def _group_by_year(folder: pathlib.Path) -> tuple[dict[int, list[pathlib.Path]], list[pathlib.Path]]:
    by_year: dict[int, list[pathlib.Path]] = {}
    fallback: list[pathlib.Path] = []
    for p in sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xlsm")):
        if p.name.startswith("~$"):
            continue
        if _is_template(p):
            continue
        yr = _year_from_filename(p)
        if yr is None:
            fallback.append(p)
        else:
            by_year.setdefault(yr, []).append(p)
    return by_year, fallback


def parse_folder_all_years(folder: pathlib.Path) -> list[ParsedSnapshot]:
    """Parse every year present in ``folder`` and return one ParsedSnapshot per
    year. Used by the worker so a backfill snapshot containing 17 years of
    files (the typical first upload) produces 17 years of dashboard data.
    """
    by_year, fallback = _group_by_year(folder)
    out: list[ParsedSnapshot] = []
    for year in sorted(by_year):
        out.append(parse_files_for_year(by_year[year], year))
    if fallback:
        from datetime import date
        out.append(parse_files_for_year(fallback, date.today().year))
    return out


def parse_folder(folder: pathlib.Path) -> ParsedSnapshot:
    """Single-year convenience for callers that only need one ParsedSnapshot
    (the dev seed, the test suite, ad-hoc scripts). Picks the latest year
    present in the folder. The RQ worker uses ``parse_folder_all_years``."""
    by_year, fallback = _group_by_year(folder)
    if not by_year:
        if not fallback:
            return ParsedSnapshot(fiscal_year=0)
        from datetime import date
        return parse_files_for_year(fallback, date.today().year)
    year = max(by_year)
    return parse_files_for_year(by_year[year], year)
