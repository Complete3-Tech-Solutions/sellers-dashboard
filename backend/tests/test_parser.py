from __future__ import annotations

import pathlib

from openpyxl import Workbook

from app.services import parser


def _write_monthly_block(ws, start_row: int, projects: list[dict]) -> int:
    """Write a single month block (header + project rows + totals) starting at start_row.
    Returns the next free row number."""
    ws.cell(row=start_row, column=1, value="###")
    ws.cell(row=start_row, column=2, value="JOB #")
    ws.cell(row=start_row, column=3, value="PROJECT NAME")
    ws.cell(row=start_row, column=4, value="% COMPL")
    ws.cell(row=start_row, column=5, value="CONTRACT")
    ws.cell(row=start_row, column=6, value="COST")
    ws.cell(row=start_row, column=7, value="PROFIT")
    ws.cell(row=start_row, column=8, value="%")
    r = start_row + 1
    total_contract = total_cost = total_profit = 0.0
    for i, p in enumerate(projects, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=p["job"])
        ws.cell(row=r, column=3, value=p["name"])
        ws.cell(row=r, column=4, value=p.get("pct_compl", 1))
        ws.cell(row=r, column=5, value=p["contract"])
        ws.cell(row=r, column=6, value=p["cost"])
        ws.cell(row=r, column=7, value=p["profit"])
        ws.cell(row=r, column=8, value=p["profit"] / p["contract"] if p["contract"] else 0)
        total_contract += p["contract"]
        total_cost += p["cost"]
        total_profit += p["profit"]
        r += 1
    # Totals row (blank in cols 1-3, numbers in 5-7)
    ws.cell(row=r, column=5, value=total_contract)
    ws.cell(row=r, column=6, value=total_cost)
    ws.cell(row=r, column=7, value=total_profit)
    return r + 1


def _write_ps_workbook(tmp: pathlib.Path, year: int, name: str | None = None) -> pathlib.Path:
    path = tmp / (name or f"{year} PS SCC.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = f"SCC {year}"
    next_row = 1
    # Two months for brevity: January + February
    next_row = _write_monthly_block(ws, next_row, [
        {"job": 100, "name": "Project Alpha", "contract": 10000, "cost": 7000, "profit": 3000},
        {"job": 101, "name": "Project Beta",  "contract": 5000,  "cost": 4000, "profit": 1000},
    ])
    next_row = _write_monthly_block(ws, next_row, [
        {"job": 102, "name": "Project Gamma", "contract": 8000, "cost": 5000, "profit": 3000},
    ])
    wb.save(path)
    return path


def test_parse_ps_workbook(tmp_path: pathlib.Path):
    _write_ps_workbook(tmp_path, 2024)
    snap = parser.parse_folder(tmp_path)
    assert snap.fiscal_year == 2024
    jobs = sorted(p.job_no for p in snap.projects)
    assert jobs == ["100", "101", "102"]
    # First two go to January, third to February
    by_job = {p.job_no: p for p in snap.projects}
    assert by_job["100"].last_month == "January"
    assert by_job["101"].last_month == "January"
    assert by_job["102"].last_month == "February"
    assert by_job["100"].profit == 3000
    # Monthly totals from blocks
    by_month = {m.month: m for m in snap.monthly}
    assert by_month["January"].gross_profit == 4000   # 3000 + 1000
    assert by_month["February"].gross_profit == 3000


def test_multiple_files_same_year_dedupe_by_mtime(tmp_path: pathlib.Path):
    """When two files map to the same year (e.g. '2015 PS SCC' + '2015 PS SCC NEW'),
    the newer file's data should win on overlapping job numbers."""
    import os
    import time

    old = _write_ps_workbook(tmp_path, 2015, name="2015 PS SCC.xlsx")
    # Bump the mtime of the OLD file backwards so the NEW file is genuinely newer.
    old_mtime = time.time() - 86400
    os.utime(old, (old_mtime, old_mtime))

    newer_path = tmp_path / "2015 PS SCC NEW.xlsx"
    wb = Workbook()
    ws = wb.active
    # Job 100 with different numbers + a new job 999
    _write_monthly_block(ws, 1, [
        {"job": 100, "name": "Updated Alpha", "contract": 99999, "cost": 1, "profit": 99998},
        {"job": 999, "name": "Project Omega", "contract": 1000, "cost": 100, "profit": 900},
    ])
    wb.save(newer_path)

    snap = parser.parse_folder(tmp_path)
    by_job = {p.job_no: p for p in snap.projects}
    # Job 100 should reflect the NEW file
    assert by_job["100"].name == "Updated Alpha"
    assert by_job["100"].contract == 99999
    # Job 999 (only in NEW) survived
    assert "999" in by_job
    # Job 101 (only in OLD) survived too
    assert "101" in by_job


def test_template_file_is_skipped(tmp_path: pathlib.Path):
    _write_ps_workbook(tmp_path, 2024, name="Profit_Summary_Template.xlsx")
    snap = parser.parse_folder(tmp_path)
    # No real year files → empty snapshot
    assert snap.projects == []


def test_year_extraction_from_real_naming():
    assert parser._year_from_filename(pathlib.Path("2013 PS SCC.xlsx")) == 2013
    assert parser._year_from_filename(pathlib.Path("2015 PS SCC NEW.xlsx")) == 2015
    assert parser._year_from_filename(pathlib.Path("2016 PS SCC - 4.22.16.xlsx")) == 2016
    assert parser._year_from_filename(pathlib.Path("2016 PS SCC 7.27.16.xlsx")) == 2016
    assert parser._year_from_filename(pathlib.Path("Profit_Summary_Template.xlsx")) is None


def test_to_number_currency_handling():
    assert parser.to_number("$1,234.56") == 1234.56
    assert parser.to_number("(500)") == -500
    assert parser.to_number("25%") == 0.25
    assert parser.to_number("") is None
    assert parser.to_number("INVOICED") is None
    assert parser.to_number("#DIV/0!") is None
    assert parser.to_number("#REF!") is None
